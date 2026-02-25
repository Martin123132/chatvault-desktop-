import json
import os
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, List, Tuple, Optional

import sqlite3
from openpyxl import load_workbook # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from chatvault_db import add_message

UPLOAD_ROOT = os.path.abspath(os.getenv("CHATVAULT_UPLOAD_DIR", "uploads"))
DEFAULT_MAX_ROWS = int(os.getenv("CHATVAULT_EXCEL_MAX_ROWS", "800"))
DEFAULT_MAX_COLS = int(os.getenv("CHATVAULT_EXCEL_MAX_COLS", "40"))
DEFAULT_MAX_CELLS = int(os.getenv("CHATVAULT_EXCEL_MAX_CELLS", "32000"))


def _coerce_value(val: Any) -> Any:
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, bytes):
        preview = val[:32]
        return preview.hex() + ("..." if len(val) > 32 else "")
    try:
        json.dumps(val)
        return val
    except TypeError:
        return str(val)


def _fetch_meta(con: sqlite3.Connection, message_id: int) -> tuple[dict[str, Any], Optional[int]]:
    cur = con.cursor()
    cur.execute("SELECT meta_json, conversation_id FROM messages WHERE id=?", (message_id,))
    row = cur.fetchone()
    if not row:
        raise ValueError(f"message_id {message_id} not found")
    raw = row[0] or "{}"
    conv_id = int(row[1]) if row[1] is not None else None
    try:
        return json.loads(raw), conv_id
    except json.JSONDecodeError:
        return {}, conv_id


def _safe_path(path: str, allow_missing: bool = False) -> str:
    # Treat bare filenames as living under the upload root so caller can pass
    # "119/myfile.xlsx" or just "myfile.xlsx" without leaking outside.
    candidate = path
    if not os.path.isabs(candidate):
        candidate = os.path.join(UPLOAD_ROOT, candidate)

    abs_path = os.path.abspath(candidate)
    if not abs_path.startswith(UPLOAD_ROOT):
        raise ValueError("file path is outside upload root")
    if not allow_missing and not os.path.exists(abs_path):
        raise ValueError("file does not exist on disk")
    return abs_path


def _find_by_basename(filename: str) -> str | None:
    """Best-effort search for a file by basename under the upload root."""
    target = filename.strip()
    if not target:
        return None
    for root, _, files in os.walk(UPLOAD_ROOT):
        if target in files:
            return os.path.abspath(os.path.join(root, target))
    return None


def resolve_file_path(con: sqlite3.Connection, file_id: str) -> Tuple[str, dict[str, Any], int | None]:
    fid = (file_id or "").strip()
    if not fid:
        raise ValueError("file_id is required")

    meta: dict[str, Any] = {}
    conv_id: int | None = None
    if fid.isdigit():
        meta, conv_id = _fetch_meta(con, int(fid))
        stored = meta.get("stored_path")
        if not stored:
            raise ValueError("message has no stored_path")
        return _safe_path(stored), meta, conv_id

    # Direct or relative path resolution.
    try:
        return _safe_path(fid), meta, conv_id
    except ValueError:
        # Fallback: search by basename anywhere under upload root for convenience.
        located = _find_by_basename(fid)
        if located:
            return _safe_path(located), meta, conv_id
        raise


def _normalize_range(ws, cell_range: str | None, max_rows: int, max_cols: int) -> Tuple[Iterable, int, int, bool, bool, bool]:
    max_rows = max(1, max_rows)
    max_cols = max(1, max_cols)
    truncated_rows = False
    truncated_cols = False
    truncated_cells = False

    if cell_range:
        selection = ws[cell_range]
        if isinstance(selection, tuple):
            if selection and isinstance(selection[0], tuple):
                rows_iter = selection
            else:
                rows_iter = (selection,)
        else:
            rows_iter = ((selection,),)
        return rows_iter, ws.max_row, ws.max_column, truncated_rows, truncated_cols, truncated_cells

    rows_cap = min(ws.max_row, max_rows)
    cols_cap = min(ws.max_column, max_cols)
    truncated_rows = ws.max_row > rows_cap
    truncated_cols = ws.max_column > cols_cap
    rows_iter = ws.iter_rows(min_row=1, max_row=rows_cap, max_col=cols_cap)
    return rows_iter, rows_cap, cols_cap, truncated_rows, truncated_cols, truncated_cells


def _cell_position(cell) -> tuple[Any, Any, Any]:
    row_idx = getattr(cell, "row", None) or getattr(cell, "row_idx", None)
    col_idx = None
    raw_col = getattr(cell, "col_idx", None)
    if raw_col is not None:
        col_idx = raw_col
    else:
        raw_col = getattr(cell, "column", None)
        if isinstance(raw_col, int):
            col_idx = raw_col
    coord = None
    if row_idx and col_idx:
        try:
            coord = f"{get_column_letter(int(col_idx))}{int(row_idx)}"
        except Exception:
            coord = None
    return row_idx, col_idx, coord


def _build_entry(cell, ws_formulas, include_values: bool, include_formulas: bool) -> dict[str, Any]:
    row_idx, col_idx, coord = _cell_position(cell)
    entry: dict[str, Any] = {
        "cell": coord,
        "row": row_idx,
        "column": col_idx,
        "type": getattr(cell, "data_type", None),
    }
    if include_values:
        entry["value"] = _coerce_value(getattr(cell, "value", None))
    if include_formulas:
        if ws_formulas is not None and coord:
            try:
                formula_cell = ws_formulas[coord]
                entry["formula"] = getattr(formula_cell, "value", None)
            except Exception:
                entry["formula"] = None
        else:
            entry["formula"] = None
    return entry


def _collect_row_cells(row, ws_formulas, include_values: bool, include_formulas: bool, consumed: int, cell_cap: int) -> Tuple[List[dict[str, Any]], int, bool]:
    row_cells: List[dict[str, Any]] = []
    for cell in row:
        if consumed >= cell_cap:
            return row_cells, consumed, True
        row_cells.append(_build_entry(cell, ws_formulas, include_values, include_formulas))
        consumed += 1
    return row_cells, consumed, False


def _serialize_rows(rows_iter: Iterable, ws_formulas, mode: str, cell_cap: int) -> Tuple[List[dict[str, Any]], bool]:
    rows: List[dict[str, Any]] = []
    consumed = 0
    include_values = mode in ("values", "both")
    include_formulas = mode in ("formulas", "both")
    truncated_cells = False

    for row in rows_iter:
        row_cells, consumed, truncated_cells = _collect_row_cells(
            row,
            ws_formulas,
            include_values,
            include_formulas,
            consumed,
            cell_cap,
        )
        row_num = None
        if row:
            row_num = getattr(row[0], "row", None) or getattr(row[0], "row_idx", None)
        rows.append({"row": row_num, "cells": row_cells})
        if truncated_cells:
            break

    return rows, truncated_cells


def inspect_excel(
    con: sqlite3.Connection,
    file_id: str,
    sheet: str | None = None,
    cell_range: str | None = None,
    mode: str = "values",
    max_rows: int | None = None,
    max_cols: int | None = None,
    max_cells: int | None = None,
) -> dict[str, Any]:
    mode = (mode or "values").lower()
    if mode not in {"values", "formulas", "both"}:
        raise ValueError("mode must be one of values, formulas, both")

    path, meta, _conv_id = resolve_file_path(con, file_id)
    ext = os.path.splitext(path)[1].lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise ValueError("only .xlsx or .xlsm files are supported")

    rows_cap = max_rows or DEFAULT_MAX_ROWS
    cols_cap = max_cols or DEFAULT_MAX_COLS
    cell_cap = max_cells or DEFAULT_MAX_CELLS

    wb_values = load_workbook(path, data_only=True, read_only=True)
    wb_formulas = None
    if mode in {"formulas", "both"}:
        wb_formulas = load_workbook(path, data_only=False, read_only=True)

    try:
        sheet_name = sheet or wb_values.sheetnames[0]
        if sheet_name not in wb_values.sheetnames:
            raise ValueError(f"sheet '{sheet_name}' not found")

        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name] if wb_formulas else None

        rows_iter, used_rows, used_cols, truncated_rows, truncated_cols, _ = _normalize_range(
            ws_values,
            cell_range,
            rows_cap,
            cols_cap,
        )
        rows_payload, truncated_cells = _serialize_rows(rows_iter, ws_formulas, mode, cell_cap)

        range_hint = cell_range
        if not range_hint:
            col_letter = get_column_letter(used_cols) if used_cols else "A"
            range_hint = f"A1:{col_letter}{used_rows}"

        return {
            "ok": True,
            "file_path": path,
            "sheet": sheet_name,
            "sheets": wb_values.sheetnames,
            "range": range_hint,
            "mode": mode,
            "dimensions": {
                "reported_rows": ws_values.max_row,
                "reported_cols": ws_values.max_column,
            },
            "truncated": {
                "rows": truncated_rows,
                "cols": truncated_cols,
                "cells": truncated_cells,
            },
            "rows": rows_payload,
            "meta": meta,
        }
    finally:
        wb_values.close()
        if wb_formulas:
            wb_formulas.close()


def _resolve_sheet(wb, raw_sheet: Any, create_sheets: bool):
    sheet_name = raw_sheet.strip() if isinstance(raw_sheet, str) and raw_sheet.strip() else wb.sheetnames[0]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name], sheet_name
    if create_sheets:
        return wb.create_sheet(sheet_name), sheet_name
    raise ValueError(f"sheet '{sheet_name}' not found and create_sheets is False")


def _normalize_cell_reference(edit: dict[str, Any]) -> str:
    raw_ref = edit.get("cell") or edit.get("range") or ""
    if isinstance(raw_ref, str):
        cell_ref = raw_ref.strip()
    elif raw_ref:
        cell_ref = str(raw_ref).strip()
    else:
        cell_ref = ""
    if not cell_ref:
        raise ValueError("each edit needs a cell or range")
    return cell_ref


def _extract_value_formula(edit: dict[str, Any]) -> Tuple[Any, Any]:
    value = edit.get("value") if "value" in edit else None
    formula = edit.get("formula") if "formula" in edit else None
    if formula is None and value is None:
        raise ValueError("each edit needs a value or formula")
    return value, formula


def _resolve_targets(ws, cell_ref: str) -> List[Any]:
    target = ws[cell_ref]
    if isinstance(target, tuple):
        return [c for row in target for c in (row if isinstance(row, (list, tuple)) else (row,))]
    return [target]


def _apply_targets(targets: List[Any], value: Any, formula: Any, sheet_name: str, applied: List[str]) -> List[str]:
    for cell in targets:
        if formula is not None:
            cell.value = formula
        else:
            cell.value = value
        applied.append(f"{sheet_name}!{cell.coordinate}")
    return applied


def modify_excel(
    con: sqlite3.Connection,
    file_id: str,
    edits: List[dict[str, Any]],
    output_mode: str = "sidecar",
    output_suffix: str = "-edited",
    create_sheets: bool = True,
) -> dict[str, Any]:
    if not edits:
        raise ValueError("edits array is required")

    path, _meta, conv_id = resolve_file_path(con, file_id)
    ext = os.path.splitext(path)[1].lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise ValueError("only .xlsx or .xlsm files are supported")

    wb = load_workbook(path)
    applied: List[str] = []

    for edit in edits:
        ws, sheet_name = _resolve_sheet(wb, edit.get("sheet"), create_sheets)
        cell_ref = _normalize_cell_reference(edit)
        value, formula = _extract_value_formula(edit)
        targets = _resolve_targets(ws, cell_ref)
        applied = _apply_targets(targets, value, formula, sheet_name, applied)

    if output_mode == "replace":
        out_path = path
    else:
        base, extension = os.path.splitext(path)
        out_path = f"{base}{output_suffix}{extension}"
        out_path = _safe_path(out_path, allow_missing=True)

    wb.save(out_path)
    wb.close()

    size_bytes = os.path.getsize(out_path)
    message_id: int | None = None
    if conv_id is not None:
        try:
            message_id = add_message(
                con=con,
                conversation_id=conv_id,
                source="ai_file",
                role="assistant",
                content=f"[AI generated Excel: {os.path.basename(out_path)} | {size_bytes} bytes | stored at {out_path}]",
                meta={
                    "filename": os.path.basename(out_path),
                    "bytes": size_bytes,
                    "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "stored_path": out_path,
                    "source": "ai_file",
                },
            )
        except Exception:
            message_id = None

    return {
        "ok": True,
        "output_path": out_path,
        "changes": applied,
        "bytes": size_bytes,
        "message_id": message_id,
    }
